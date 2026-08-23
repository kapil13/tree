"""Compliance Phase E — science & buyer differentiation."""

from __future__ import annotations

import io
import uuid
import zipfile
from unittest.mock import AsyncMock, MagicMock

import pytest
from openpyxl import load_workbook

from app.services.biodiversity.darwin_core import render_darwin_zip
from app.services.compliance.checklists import CHECKLISTS
from app.services.reports.eudr_exports import (
    build_eudr_due_diligence_context,
    render_eudr_due_diligence_xlsx,
)
from app.services.reports.gbf_exports import build_gbf_context, build_gbf_tnfd_section, render_gbf_xlsx
from app.services.reports.iso14064_org import build_iso14064_org_context, render_iso14064_org_xlsx
from app.services.reports.sbti_flag import build_sbti_flag_context, render_sbti_flag_xlsx


def test_sbti_flag_and_eudr_checklists_registered():
    assert "sbti_flag" in CHECKLISTS
    assert "eudr_supplier_mrv" in CHECKLISTS
    flag_auto = {i.auto_key for i in CHECKLISTS["sbti_flag"].items if i.auto_key}
    assert "flag_land_boundary" in flag_auto
    assert "leakage_documented" in flag_auto
    eudr_auto = {i.auto_key for i in CHECKLISTS["eudr_supplier_mrv"].items if i.auto_key}
    assert "eudr_geo_due_diligence" in eudr_auto
    assert "supplier_ref_documented" in eudr_auto


@pytest.mark.asyncio
async def test_build_sbti_flag_context(monkeypatch):
    org_id = uuid.uuid4()
    project = MagicMock()
    project.id = uuid.uuid4()
    project.code = "P-FLAG"
    project.name = "Flag Demo"
    project.organization_id = org_id
    project.metadata_ = {}

    db = AsyncMock()
    db.execute = AsyncMock(
        return_value=MagicMock(scalars=MagicMock(return_value=MagicMock(all=lambda: [project])))
    )

    async def fake_row(db_, proj):
        return {
            "project_code": proj.code,
            "project_name": proj.name,
            "reporting_year": 2026,
            "land_area_ha": 12.5,
            "tree_count": 100,
            "geo_tagged_pct": 85.0,
            "gross_removals_tco2e": 10.0,
            "leakage_tco2e": 0.5,
            "net_land_removals_tco2e": 9.5,
            "buffer_pct": 0.2,
            "flag_category": "land_removals_arr",
            "data_source": "vm0047_ghg_single_source",
            "engine_version": "test",
            "inventory_net_tco2e": 8.0,
        }

    monkeypatch.setattr("app.services.reports.sbti_flag._project_flag_row", fake_row)

    ctx = await build_sbti_flag_context(db, org_id)
    assert ctx["export_type"] == "sbti_flag_worksheet"
    assert ctx["totals"]["project_count"] == 1
    assert ctx["rows"][0]["net_land_removals_tco2e"] == 9.5
    xlsx = render_sbti_flag_xlsx(ctx)
    wb = load_workbook(io.BytesIO(xlsx))
    assert "SBTi FLAG" in wb.sheetnames


@pytest.mark.asyncio
async def test_build_eudr_due_diligence_context(monkeypatch):
    project = MagicMock()
    project.id = uuid.uuid4()
    project.code = "P-EUDR"
    project.name = "EUDR Demo"
    project.scheme_code = "campa_ca"
    project.segment = "govt"
    project.metadata_ = {"scheme_refs": {"supplier_ref": "SUP-001", "state_name": "Maharashtra"}}

    db = AsyncMock()

    async def fake_mrv(db_, proj):
        return {"summary": {"open_violations": 0, "native_species_pct": 80}}

    async def fake_trees(db_):
        return []

    async def fake_fences(db_):
        return MagicMock(scalars=MagicMock(return_value=MagicMock(all=lambda: [])))

    async def fake_safeguards(db_, pid):
        return []

    monkeypatch.setattr(
        "app.services.reports.eudr_exports.build_project_mrv_context", fake_mrv
    )
    monkeypatch.setattr(
        "app.services.reports.eudr_exports.list_safeguard_documents", fake_safeguards
    )

    call_count = {"n": 0}

    async def fake_execute(stmt):
        call_count["n"] += 1
        if call_count["n"] == 1:
            return MagicMock(scalars=MagicMock(return_value=MagicMock(all=lambda: [])))
        return MagicMock(scalars=MagicMock(return_value=MagicMock(all=lambda: [])))

    db.execute = fake_execute

    ctx = await build_eudr_due_diligence_context(db, project)
    assert ctx["supplier"]["supplier_ref"] == "SUP-001"
    assert ctx["brsr_annex"][0]["supplier_ref"] == "SUP-001"
    xlsx = render_eudr_due_diligence_xlsx(ctx)
    assert len(xlsx) > 500


@pytest.mark.asyncio
async def test_build_gbf_context_and_tnfd_bridge(monkeypatch):
    org = MagicMock()
    org.id = uuid.uuid4()
    org.name = "Demo Org"
    org.slug = "demo"

    project = MagicMock()
    project.id = uuid.uuid4()
    project.code = "P-GBF"
    project.name = "GBF Demo"

    db = AsyncMock()
    db.execute = AsyncMock(
        return_value=MagicMock(scalars=MagicMock(return_value=MagicMock(all=lambda: [project])))
    )

    async def fake_mrv(db_, proj):
        return {"summary": {"tree_count": 50, "survival_pct": 90, "native_species_pct": 75, "open_violations": 0}}

    async def fake_fences(db_):
        return MagicMock(scalars=MagicMock(return_value=MagicMock(all=lambda: [])))

    call_count = {"n": 0}

    async def fake_execute(stmt):
        call_count["n"] += 1
        if call_count["n"] == 1:
            return MagicMock(scalars=MagicMock(return_value=MagicMock(all=lambda: [project])))
        return MagicMock(scalars=MagicMock(return_value=MagicMock(all=lambda: [])))

    db.execute = fake_execute
    monkeypatch.setattr("app.services.reports.gbf_exports.build_project_mrv_context", fake_mrv)

    ctx = await build_gbf_context(db, organization=org)
    assert ctx["target_2_restore"]["portfolio_totals"]["trees_planted"] == 50
    bridge = build_gbf_tnfd_section(ctx)
    assert bridge["framework"] == "Kunming-Montreal GBF"
    assert render_gbf_xlsx(ctx)


@pytest.mark.asyncio
async def test_iso14064_org_context(monkeypatch):
    org = MagicMock()
    org.id = uuid.uuid4()
    org.name = "Demo Org"
    org.slug = "demo"

    db = AsyncMock()

    async def fake_ghg(db_, **kwargs):
        return {
            "organization": {"id": str(org.id), "name": org.name, "slug": org.slug},
            "reporting_boundary": "Operational control",
            "inventory_lines": [
                {
                    "line_id": "P1-LSR-GROSS",
                    "project_code": "P1",
                    "ghg_protocol_category": "Land Sector Removals",
                    "gas": "CO2",
                    "amount_tco2e": 5.0,
                    "uncertainty_pct": 20,
                }
            ],
            "portfolio_summary": {"project_count": 1, "total_gross_removals_tco2e": 5.0},
        }

    monkeypatch.setattr("app.services.reports.iso14064_org.build_ghg_protocol_context", fake_ghg)

    ctx = await build_iso14064_org_context(db, organization=org)
    assert ctx["standard"] == "ISO 14064-1"
    assert any(line.get("data_available") for line in ctx["inventory_lines"])
    xlsx = render_iso14064_org_xlsx(ctx)
    assert len(xlsx) > 500


def test_darwin_zip_includes_gbif_prep_and_iucn_fields():
    occurrences = [
        {
            "occurrenceID": "test-1",
            "scientificName": "Paradoxurus hermaphroditus",
            "iucnRedListCategory": "LC",
            "iucnTaxonID": "12345",
            "gbifID": 987654,
        }
    ]
    zbytes = render_darwin_zip(occurrences, project_code="P1", org_name="Demo")
    with zipfile.ZipFile(io.BytesIO(zbytes)) as zf:
        names = zf.namelist()
        assert "gbif_publish_prep.json" in names
        assert "occurrence.txt" in names
        tsv = zf.read("occurrence.txt").decode("utf-8")
        assert "iucnTaxonID" in tsv
        assert "gbifID" in tsv
