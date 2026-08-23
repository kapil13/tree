"""India portal-shaped compliance exports (Phase A — CAMPA, Green Credit)."""

from __future__ import annotations

import io
from datetime import UTC, datetime
from typing import Any

from openpyxl import Workbook
from sqlalchemy.ext.asyncio import AsyncSession

from app.models.planting_project import PlantingProject
from app.services.credits.green_credit import build_project_green_credit_summary
from app.services.planting_projects.mrv_export import build_project_mrv_context
from app.services.schemes.kpis import compute_scheme_kpis

DISCLAIMER = (
    "Prepared for manual upload to state/MoEFCC portals. Not an official registry submission."
)


async def build_campa_state_export_context(
    db: AsyncSession, project: PlantingProject
) -> dict[str, Any]:
    mrv = await build_project_mrv_context(db, project)
    kpis = await compute_scheme_kpis(db, project)
    refs = (project.metadata_ or {}).get("scheme_refs") or {}
    summary = mrv.get("summary") or {}
    tree_count = int(summary.get("tree_count") or 0)
    geo_tagged = int(summary.get("geo_tagged_count") or 0)
    geo_pct = round(100 * geo_tagged / tree_count, 1) if tree_count else 0.0
    survival_pct = float((kpis.get("metrics") or {}).get("survival_pct") or 0)

    return {
        "export_type": "state_campa_monitoring",
        "generated_at": datetime.now(UTC).isoformat(),
        "disclaimer": DISCLAIMER,
        "project": {
            "code": project.code,
            "name": project.name,
            "scheme_code": project.scheme_code,
            "segment": project.segment,
            "status": project.status,
        },
        "scheme_refs": refs,
        "monitoring": {
            "tree_count": tree_count,
            "geo_tagged_count": geo_tagged,
            "geo_tagged_pct": geo_pct,
            "survival_pct": survival_pct,
            "open_violations": int(summary.get("open_violations") or 0),
            "blocking_violations": int(summary.get("blocking_violations") or 0),
            "work_area_count": int(summary.get("work_area_count") or 0),
        },
        "kpi_status": kpis.get("status"),
        "kpi_targets": kpis.get("targets"),
        "portal_row": {
            "state_ut": refs.get("state_name", ""),
            "apo_financial_year": refs.get("apo_financial_year", ""),
            "pca_number": refs.get("pca_number", ""),
            "forest_diversion_id": refs.get("forest_diversion_id", ""),
            "state_campa_account": refs.get("state_campa_account", ""),
            "ca_land_parcel_id": refs.get("ca_land_parcel_id", ""),
            "ngt_case_number": refs.get("ngt_case_number", ""),
            "ngt_order_ref": refs.get("ngt_order_ref", ""),
            "geo_tagged_pct": geo_pct,
            "survival_pct": survival_pct,
            "violation_status": "clear"
            if int(summary.get("blocking_violations") or 0) == 0
            else "blocking_open",
        },
    }


def render_campa_state_export_xlsx(ctx: dict[str, Any]) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "State CAMPA"
    row = ctx.get("portal_row") or {}
    headers = [
        "State/UT",
        "APO FY",
        "PCA number",
        "FC diversion ref",
        "State CAMPA account",
        "CA land parcel",
        "NGT case",
        "NGT order ref",
        "Geo-tagged %",
        "Survival %",
        "Violation status",
        "Project code",
        "Tree count",
    ]
    ws.append(headers)
    monitoring = ctx.get("monitoring") or {}
    ws.append(
        [
            row.get("state_ut"),
            row.get("apo_financial_year"),
            row.get("pca_number"),
            row.get("forest_diversion_id"),
            row.get("state_campa_account"),
            row.get("ca_land_parcel_id"),
            row.get("ngt_case_number"),
            row.get("ngt_order_ref"),
            row.get("geo_tagged_pct"),
            row.get("survival_pct"),
            row.get("violation_status"),
            (ctx.get("project") or {}).get("code"),
            monitoring.get("tree_count"),
        ]
    )
    ws2 = wb.create_sheet("Disclaimer")
    ws2.append(["Disclaimer"])
    ws2.append([ctx.get("disclaimer", "")])
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


async def build_green_credit_portal_context(
    db: AsyncSession, project: PlantingProject
) -> dict[str, Any]:
    summary = await build_project_green_credit_summary(db, project)
    refs = (project.metadata_ or {}).get("scheme_refs") or {}
    return {
        "export_type": "green_credit_registrar_handoff",
        "generated_at": datetime.now(UTC).isoformat(),
        "disclaimer": summary.get("disclaimer", DISCLAIMER),
        "registrar_columns": {
            "land_bank_id": summary.get("land_bank_id") or refs.get("green_credit_land_bank_id"),
            "activity_type": summary.get("activity_type"),
            "project_code": summary.get("project_code"),
            "verifier_reference": summary.get("verifier_reference")
            or refs.get("verifier_reference"),
            "tree_count": summary.get("tree_count"),
            "eligible_trees": summary.get("eligible_trees"),
            "total_area_ha": summary.get("total_area_ha"),
            "trees_per_ha": summary.get("trees_per_ha"),
            "min_trees_per_ha": summary.get("min_trees_per_ha"),
            "density_eligible": summary.get("density_eligible"),
            "survival_pct": summary.get("survival_pct_assumed"),
            "vested_green_credits": summary.get("vested_green_credits"),
            "provisional_green_credits": summary.get("provisional_green_credits"),
            "eligibility_status": summary.get("eligibility_status"),
            "gaps": ", ".join(summary.get("gaps") or []),
        },
        "summary": summary,
    }


def render_green_credit_portal_xlsx(ctx: dict[str, Any]) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Green Credit handoff"
    cols = ctx.get("registrar_columns") or {}
    ws.append(["Field", "Value"])
    for key, value in cols.items():
        ws.append([key, value])
    ws2 = wb.create_sheet("Disclaimer")
    ws2.append([ctx.get("disclaimer", "")])
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
