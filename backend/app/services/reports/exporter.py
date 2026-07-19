"""PDF and Excel report exporters."""

from __future__ import annotations

import io
from datetime import datetime
from typing import Any

from openpyxl import Workbook
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.lib.units import mm
from reportlab.platypus import (
    Paragraph,
    SimpleDocTemplate,
    Spacer,
    Table,
    TableStyle,
)


def render_carbon_report_pdf(
    org_name: str, summary: dict[str, Any], rows: list[dict[str, Any]]
) -> bytes:
    buf = io.BytesIO()
    doc = SimpleDocTemplate(
        buf,
        pagesize=A4,
        title=f"BYOT Carbon Report - {org_name}",
        author="BYOT",
        leftMargin=18 * mm,
        rightMargin=18 * mm,
        topMargin=18 * mm,
        bottomMargin=18 * mm,
    )
    styles = getSampleStyleSheet()
    h1 = styles["Heading1"]
    h1.textColor = colors.HexColor("#15803D")
    body = styles["BodyText"]

    story: list = []
    story.append(Paragraph("Carbon Sequestration Report", h1))
    story.append(
        Paragraph(
            f"<b>{org_name}</b> · generated {datetime.utcnow().strftime('%Y-%m-%d %H:%M UTC')}",
            body,
        )
    )
    story.append(Spacer(1, 6 * mm))

    summary_rows = [
        ["Total trees", summary.get("total_trees", 0)],
        ["Total biomass (kg)", f"{summary.get('total_biomass_kg', 0):,.2f}"],
        ["Total carbon (kg C)", f"{summary.get('total_carbon_kg', 0):,.2f}"],
        ["Total CO2e (kg)", f"{summary.get('total_co2e_kg', 0):,.2f}"],
        ["Annual sequestration (kg CO2e)", f"{summary.get('annual_sequestration_kg', 0):,.2f}"],
        ["Lifetime credits (tCO2e)", f"{summary.get('lifetime_credits_tco2e', 0):,.3f}"],
        ["Estimated revenue (USD)", f"${summary.get('estimated_revenue_usd', 0):,.2f}"],
    ]
    t = Table(summary_rows, colWidths=[80 * mm, 80 * mm])
    t.setStyle(
        TableStyle(
            [
                ("FONTNAME", (0, 0), (-1, -1), "Helvetica"),
                ("BACKGROUND", (0, 0), (0, -1), colors.HexColor("#DCFCE7")),
                ("TEXTCOLOR", (0, 0), (0, -1), colors.HexColor("#15803D")),
                ("LINEBELOW", (0, 0), (-1, -1), 0.25, colors.grey),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 6),
            ]
        )
    )
    story.append(t)
    story.append(Spacer(1, 8 * mm))

    story.append(Paragraph("Top trees by sequestration", body))
    detail = [["Public code", "Species", "Health", "Carbon (kg)", "CO2e (kg)"]]
    for r in rows[:25]:
        detail.append(
            [
                r.get("public_code", ""),
                r.get("species", ""),
                r.get("health", ""),
                f"{r.get('carbon_kg', 0):.2f}",
                f"{r.get('co2e_kg', 0):.2f}",
            ]
        )
    dt = Table(detail, repeatRows=1)
    dt.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#15803D")),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                ("FONTNAME", (0, 0), (-1, -1), "Helvetica"),
                ("FONTSIZE", (0, 0), (-1, -1), 9),
                ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.whitesmoke, colors.white]),
                ("GRID", (0, 0), (-1, -1), 0.1, colors.grey),
            ]
        )
    )
    story.append(dt)
    doc.build(story)
    return buf.getvalue()


def render_trees_report_xlsx(rows: list[dict[str, Any]]) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Trees"
    headers = [
        "public_code",
        "species",
        "health",
        "lat",
        "lon",
        "planted_at",
        "dbh_cm",
        "height_m",
        "carbon_kg",
        "co2e_kg",
        "satellite_verified",
    ]
    ws.append(headers)
    for r in rows:
        ws.append([r.get(h) for h in headers])
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def render_bioacoustic_report_pdf(
    title: str, ecosystem: dict[str, Any], recordings: list[dict[str, Any]]
) -> bytes:
    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=A4, title=title, author="BYOT")
    styles = getSampleStyleSheet()
    h1 = styles["Heading1"]
    h1.textColor = colors.HexColor("#15803D")
    body = styles["BodyText"]
    story: list = []

    story.append(Paragraph("Biodiversity & Bioacoustic Report", h1))
    story.append(
        Paragraph(
            f"<b>{title}</b> · {datetime.utcnow().strftime('%Y-%m-%d %H:%M UTC')}",
            body,
        )
    )
    story.append(Spacer(1, 5 * mm))

    bio = ecosystem.get("bioacoustic") or ecosystem
    summary_rows = [
        ["Recordings analyzed", bio.get("recording_count", 0)],
        ["Bioacoustic health (avg)", f"{bio.get('avg_health_score', 0)}/100"],
        ["Shannon H′ (avg)", str(bio.get("avg_shannon_index", 0))],
        ["Simpson D (avg)", str(bio.get("avg_simpson_index", 0))],
        ["Species detected", bio.get("total_species_detected", 0)],
        ["Threatened detections", bio.get("threatened_species_count", 0)],
        ["NDVI mean", f"{ecosystem.get('ndvi_mean', '—')}"],
        ["NDVI trend", ecosystem.get("ndvi_trend") or "—"],
        ["Ecosystem score", f"{ecosystem.get('ecosystem_health_score', '—')}/100"],
    ]
    t = Table(summary_rows, colWidths=[75 * mm, 85 * mm])
    t.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (0, -1), colors.HexColor("#DCFCE7")),
                ("GRID", (0, 0), (-1, -1), 0.1, colors.grey),
                ("FONTSIZE", (0, 0), (-1, -1), 9),
            ]
        )
    )
    story.append(t)
    story.append(Spacer(1, 5 * mm))

    if ecosystem.get("interpretation"):
        story.append(Paragraph("<b>Interpretation</b>", body))
        story.append(Paragraph(ecosystem["interpretation"], body))
        story.append(Spacer(1, 5 * mm))

    taxon = bio.get("taxon_breakdown") or {}
    if taxon:
        story.append(Paragraph("<b>Taxon activity (call counts)</b>", body))
        taxon_rows = [["Taxon group", "Calls"]] + [[k, v] for k, v in sorted(taxon.items())]
        story.append(Table(taxon_rows, colWidths=[60 * mm, 40 * mm]))
        story.append(Spacer(1, 5 * mm))

    story.append(Paragraph("<b>Top species</b>", body))
    species = bio.get("species_list") or []
    sp_table = [["Common name", "Scientific", "Taxon", "Calls", "IUCN"]]
    for s in species[:15]:
        sp_table.append(
            [
                s.get("common_name", ""),
                s.get("scientific_name", ""),
                s.get("taxon_group", ""),
                str(s.get("call_count", 0)),
                s.get("iucn_status", ""),
            ]
        )
    st = Table(sp_table, repeatRows=1, colWidths=[35 * mm, 45 * mm, 22 * mm, 18 * mm, 30 * mm])
    st.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#15803D")),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                ("FONTSIZE", (0, 0), (-1, -1), 8),
                ("GRID", (0, 0), (-1, -1), 0.1, colors.grey),
            ]
        )
    )
    story.append(st)

    if recordings:
        story.append(Spacer(1, 5 * mm))
        story.append(Paragraph("<b>Recent recordings</b>", body))
        rec_table = [["Date", "Duration", "Health", "Species"]]
        for r in recordings[:10]:
            rec_table.append(
                [
                    str(r.get("recorded_at", ""))[:16],
                    f"{r.get('duration_seconds', 0)}s",
                    str(r.get("bioacoustic_health_score", "—")),
                    str(r.get("total_species_count", "—")),
                ]
            )
        story.append(Table(rec_table, repeatRows=1))

    doc.build(story)
    return buf.getvalue()


def render_esg_report_pdf(
    org_name: str,
    carbon_summary: dict[str, Any],
    ecosystem: dict[str, Any] | None,
    tree_rows: list[dict[str, Any]],
) -> bytes:
    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=A4, title=f"ESG - {org_name}", author="BYOT")
    styles = getSampleStyleSheet()
    h1 = styles["Heading1"]
    h1.textColor = colors.HexColor("#15803D")
    body = styles["BodyText"]
    story: list = []

    story.append(Paragraph("ESG Impact Report", h1))
    story.append(
        Paragraph(
            f"<b>{org_name}</b> · Carbon + Biodiversity + Satellite · "
            f"{datetime.utcnow().strftime('%Y-%m-%d %H:%M UTC')}",
            body,
        )
    )
    story.append(Spacer(1, 5 * mm))

    story.append(Paragraph("<b>Carbon & climate</b>", body))
    carbon_rows = [
        ["Total trees", carbon_summary.get("total_trees", 0)],
        ["Total CO2e (kg)", f"{carbon_summary.get('total_co2e_kg', 0):,.2f}"],
        ["Lifetime credits (tCO2e)", f"{carbon_summary.get('lifetime_credits_tco2e', 0):,.3f}"],
        ["Est. revenue (USD)", f"${carbon_summary.get('estimated_revenue_usd', 0):,.2f}"],
    ]
    story.append(Table(carbon_rows, colWidths=[80 * mm, 80 * mm]))
    story.append(Spacer(1, 6 * mm))

    if ecosystem:
        story.append(Paragraph("<b>Biodiversity & ecosystem health</b>", body))
        bio = ecosystem.get("bioacoustic") or {}
        eco_rows = [
            ["Site", ecosystem.get("fence_name", org_name)],
            ["Bioacoustic health", f"{bio.get('avg_health_score', 0)}/100"],
            ["Species detected", bio.get("total_species_detected", 0)],
            ["NDVI", ecosystem.get("ndvi_mean")],
            ["Correlation (bio × NDVI)", ecosystem.get("correlation_score")],
            ["Ecosystem score", f"{ecosystem.get('ecosystem_health_score', 0)}/100"],
        ]
        story.append(Table(eco_rows, colWidths=[80 * mm, 80 * mm]))
        if ecosystem.get("interpretation"):
            story.append(Spacer(1, 3 * mm))
            story.append(Paragraph(ecosystem["interpretation"], body))

    story.append(Spacer(1, 6 * mm))
    story.append(Paragraph("<b>Top trees</b>", body))
    detail = [["Code", "Species", "Carbon kg", "Health"]]
    for r in tree_rows[:15]:
        detail.append(
            [
                r.get("public_code", ""),
                r.get("species", ""),
                f"{r.get('carbon_kg', 0):.1f}",
                r.get("health", ""),
            ]
        )
    story.append(Table(detail, repeatRows=1))

    doc.build(story)
    return buf.getvalue()


def render_bioacoustic_report_xlsx(ecosystem: dict[str, Any]) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Biodiversity"
    bio = ecosystem.get("bioacoustic") or ecosystem
    ws.append(["metric", "value"])
    for key in (
        "recording_count",
        "avg_health_score",
        "avg_shannon_index",
        "avg_simpson_index",
        "total_species_detected",
        "threatened_species_count",
    ):
        ws.append([key, bio.get(key)])
    ws.append([])
    ws.append(["scientific_name", "common_name", "taxon_group", "call_count", "iucn_status"])
    for s in bio.get("species_list") or []:
        ws.append(
            [
                s.get("scientific_name"),
                s.get("common_name"),
                s.get("taxon_group"),
                s.get("call_count"),
                s.get("iucn_status"),
            ]
        )
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def render_compliance_mrv_pdf(ctx: dict[str, Any]) -> bytes:
    """Project-level MRV / compliance audit PDF."""
    buf = io.BytesIO()
    project = ctx["project"]
    summary = ctx["summary"]
    doc = SimpleDocTemplate(
        buf,
        pagesize=A4,
        title=f"MRV Compliance - {project['name']}",
        author="Aranyix BYOT",
        leftMargin=18 * mm,
        rightMargin=18 * mm,
        topMargin=18 * mm,
        bottomMargin=18 * mm,
    )
    styles = getSampleStyleSheet()
    h1 = styles["Heading1"]
    h1.textColor = colors.HexColor("#15803D")
    h2 = styles["Heading2"]
    body = styles["BodyText"]

    story: list = []
    story.append(Paragraph("MRV Compliance Report", h1))
    story.append(
        Paragraph(
            f"<b>{project['name']}</b> ({project['code']}) · "
            f"{project['segment'].upper()} · mode: {project['compliance_mode']} · "
            f"generated {datetime.utcnow().strftime('%Y-%m-%d %H:%M UTC')}",
            body,
        )
    )
    if project.get("standard_name"):
        story.append(
            Paragraph(
                f"Standard: {project['standard_name']} ({project.get('standard_template') or 'custom'})",
                body,
            )
        )
    story.append(Spacer(1, 6 * mm))

    kpi_rows = [
        ["Work areas", summary.get("work_area_count", 0)],
        ["Trees registered", summary.get("tree_count", 0)],
        ["Open violations", summary.get("open_violations", 0)],
        ["Resolved violations", summary.get("resolved_violations", 0)],
        ["Native species %", summary.get("native_species_pct") or "—"],
    ]
    t = Table(kpi_rows, colWidths=[80 * mm, 80 * mm])
    t.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (0, -1), colors.HexColor("#DCFCE7")),
                ("TEXTCOLOR", (0, 0), (0, -1), colors.HexColor("#15803D")),
                ("GRID", (0, 0), (-1, -1), 0.25, colors.grey),
            ]
        )
    )
    story.append(t)
    story.append(Spacer(1, 6 * mm))

    rules = ctx.get("rules_summary") or {}
    if any(rules.values()):
        story.append(Paragraph("Active compliance rules", h2))
        rule_rows = [[k.replace("_", " "), str(v)] for k, v in rules.items() if v is not None]
        if rule_rows:
            story.append(Table(rule_rows, colWidths=[80 * mm, 80 * mm]))
            story.append(Spacer(1, 6 * mm))

    survival = summary.get("survival_counts") or {}
    if survival:
        story.append(Paragraph("Survival survey status", h2))
        story.append(Table([[s, c] for s, c in sorted(survival.items())], colWidths=[80 * mm, 80 * mm]))
        story.append(Spacer(1, 6 * mm))

    story.append(Paragraph("Work areas", h2))
    wa_table = [["Name", "Type", "Area (ha)", "Trees", "Chainage"]]
    for w in ctx.get("work_areas") or []:
        chain = ""
        if w.get("chainage_start_km") is not None:
            chain = f"{w['chainage_start_km']}–{w.get('chainage_end_km', '')} km"
        wa_table.append(
            [
                w.get("name", ""),
                w.get("geometry_type", ""),
                str(w.get("area_ha") or "—"),
                str(w.get("tree_count", 0)),
                chain,
            ]
        )
    story.append(Table(wa_table, repeatRows=1))
    story.append(Spacer(1, 6 * mm))

    story.append(Paragraph("Compliance violations (recent)", h2))
    v_table = [["Severity", "Type", "Message", "Resolved"]]
    for v in ctx.get("violations") or []:
        v_table.append(
            [
                v.get("severity", ""),
                v.get("violation_type", ""),
                (v.get("message") or "")[:80],
                "yes" if v.get("resolved") else "no",
            ]
        )
    if len(v_table) == 1:
        v_table.append(["—", "—", "No violations recorded", "—"])
    story.append(Table(v_table, repeatRows=1))
    story.append(Spacer(1, 6 * mm))

    story.append(Paragraph("Tree registry (sample)", h2))
    tree_table = [["Code", "Species", "Health", "Survival", "Chainage"]]
    for tr in (ctx.get("trees") or [])[:40]:
        tree_table.append(
            [
                tr.get("public_code", ""),
                tr.get("species", ""),
                tr.get("health", ""),
                tr.get("survival_status", ""),
                str(tr.get("chainage_km") or "—"),
            ]
        )
    story.append(Table(tree_table, repeatRows=1))

    doc.build(story)
    return buf.getvalue()


def render_compliance_mrv_xlsx(ctx: dict[str, Any]) -> bytes:
    wb = Workbook()
    project = ctx["project"]
    summary = ctx["summary"]

    ws = wb.active
    ws.title = "Summary"
    ws.append(["field", "value"])
    ws.append(["project_code", project.get("code")])
    ws.append(["project_name", project.get("name")])
    ws.append(["segment", project.get("segment")])
    ws.append(["compliance_mode", project.get("compliance_mode")])
    ws.append(["trees", summary.get("tree_count")])
    ws.append(["open_violations", summary.get("open_violations")])
    ws.append(["resolved_violations", summary.get("resolved_violations")])
    ws.append(["native_species_pct", summary.get("native_species_pct")])

    ws2 = wb.create_sheet("Work areas")
    ws2.append(
        ["name", "geometry_type", "area_ha", "tree_count", "segment_code", "chainage_start_km", "chainage_end_km"]
    )
    for w in ctx.get("work_areas") or []:
        ws2.append(
            [
                w.get("name"),
                w.get("geometry_type"),
                w.get("area_ha"),
                w.get("tree_count"),
                w.get("segment_code"),
                w.get("chainage_start_km"),
                w.get("chainage_end_km"),
            ]
        )

    ws3 = wb.create_sheet("Trees")
    ws3.append(
        ["public_code", "species", "health", "survival_status", "chainage_km", "lat", "lon", "planted_at", "last_geotag_at"]
    )
    for tr in ctx.get("trees") or []:
        ws3.append(
            [
                tr.get("public_code"),
                tr.get("species"),
                tr.get("health"),
                tr.get("survival_status"),
                tr.get("chainage_km"),
                tr.get("lat"),
                tr.get("lon"),
                tr.get("planted_at"),
                tr.get("last_geotag_at"),
            ]
        )

    ws4 = wb.create_sheet("Violations")
    ws4.append(["severity", "violation_type", "message", "tree_id", "resolved", "created_at"])
    for v in ctx.get("violations") or []:
        ws4.append(
            [
                v.get("severity"),
                v.get("violation_type"),
                v.get("message"),
                v.get("tree_id"),
                v.get("resolved"),
                v.get("created_at"),
            ]
        )

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
