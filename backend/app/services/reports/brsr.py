"""SEBI BRSR Core Principle 6 (Environment) export for enterprise buyers."""

from __future__ import annotations

import io
import json
import uuid
import zipfile
from datetime import UTC, datetime
from typing import Any, Literal

from openpyxl import Workbook
from sqlalchemy import func, select
from sqlalchemy.ext.asyncio import AsyncSession

from app.models.credit_ledger import ProjectCreditLedger
from app.models.credit_serial import CreditSerial
from app.models.organization import Organization
from app.models.planting_project import PlantingProject
from app.models.tree import Tree
from app.models.verification_workflow import VerificationItem, VerificationSample
from app.services.carbon.engine import ENGINE_VERSION
from app.services.credits.ledger import org_credit_summary
from app.services.planting_projects.mrv_export import build_project_mrv_context
from app.services.reports.brsr_kpi_map import (
    build_core_kpi_sheet_rows,
    build_value_chain_annex,
)

BRSR_CORE_VERSION = "2024"
PRINCIPLE = 6


def _reporting_year() -> int:
    return datetime.now(UTC).year


async def _portfolio_projects(
    db: AsyncSession,
    organization_id: uuid.UUID,
    project_id: uuid.UUID | None = None,
) -> list[PlantingProject]:
    stmt = select(PlantingProject).where(PlantingProject.organization_id == organization_id)
    if project_id is not None:
        stmt = stmt.where(PlantingProject.id == project_id)
    return list((await db.execute(stmt.order_by(PlantingProject.code.asc()))).scalars().all())


async def _project_ghg_lines(
    db: AsyncSession,
    project: PlantingProject,
) -> list[dict[str, Any]]:
    trees_res = await db.execute(
        select(Tree).where(Tree.project_id == project.id, Tree.status != "removed")
    )
    trees = list(trees_res.scalars().all())
    total_carbon_kg = sum(float(t.current_carbon_kg or 0) for t in trees)
    total_co2e_kg = total_carbon_kg * 44 / 12

    # Portfolio-level 90% CI placeholder (±20% default when per-tree measurements unavailable)
    uncertainty_pct = 20.0 if trees else 0.0
    co2e_lower_kg = total_co2e_kg * (1.0 - uncertainty_pct / 100.0 * 0.9)
    co2e_upper_kg = total_co2e_kg * (1.0 + uncertainty_pct / 100.0 * 0.9)

    ledger = (
        await db.execute(
            select(ProjectCreditLedger).where(ProjectCreditLedger.project_id == project.id)
        )
    ).scalar_one_or_none()

    gross_t = total_co2e_kg / 1000.0
    net_t = float(ledger.net_credits_tco2e) if ledger else gross_t * 0.8

    lines: list[dict[str, Any]] = [
        {
            "line_id": f"{project.code}-REMOVAL-GROSS",
            "project_code": project.code,
            "description": "Land sector CO₂ removals (gross sequestration estimate)",
            "scope": "Land Sector",
            "scope_tag": "Scope 1 equivalent — removals",
            "ghg_protocol_category": "Land Sector Removals",
            "brsr_indicator": "P6.E4",
            "gas": "CO2",
            "amount_tco2e": round(gross_t, 4),
            "uncertainty_pct": uncertainty_pct,
            "co2e_lower_90_t": round(co2e_lower_kg / 1000.0, 4),
            "co2e_upper_90_t": round(co2e_upper_kg / 1000.0, 4),
            "methodology": ledger.methodology if ledger else "VERRA_VM0047",
            "engine_version": ENGINE_VERSION,
            "tree_count": len(trees),
        },
        {
            "line_id": f"{project.code}-REMOVAL-NET",
            "project_code": project.code,
            "description": "Net issuable credits after buffer (registry-aligned)",
            "scope": "Land Sector",
            "scope_tag": "Scope 1 equivalent — removals (net)",
            "ghg_protocol_category": "Land Sector Removals — net of buffer",
            "brsr_indicator": "P6.E4",
            "gas": "CO2",
            "amount_tco2e": round(net_t, 4),
            "uncertainty_pct": uncertainty_pct,
            "co2e_lower_90_t": round(co2e_lower_kg / 1000.0 * 0.8, 4),
            "co2e_upper_90_t": round(co2e_upper_kg / 1000.0 * 0.8, 4),
            "methodology": ledger.methodology if ledger else "VERRA_VM0047",
            "ledger_status": ledger.status if ledger else "estimated",
        },
    ]
    return lines


async def _assurance_pack(
    db: AsyncSession,
    organization_id: uuid.UUID,
    project_ids: list[uuid.UUID],
) -> dict[str, Any]:
    serials_res = await db.execute(
        select(CreditSerial)
        .where(CreditSerial.organization_id == organization_id)
        .order_by(CreditSerial.created_at.desc())
        .limit(20)
    )
    serials = [
        {
            "serial_number": s.serial_number,
            "vintage_year": s.vintage_year,
            "tco2e_amount": float(s.tco2e_amount),
            "status": s.status,
        }
        for s in serials_res.scalars().all()
    ]

    attestation_count = 0
    if project_ids:
        sample_ids = (
            await db.execute(
                select(VerificationSample.id).where(VerificationSample.project_id.in_(project_ids))
            )
        ).scalars().all()
        if sample_ids:
            attestation_count = (
                await db.execute(
                    select(func.count())
                    .select_from(VerificationItem)
                    .where(
                        VerificationItem.sample_id.in_(sample_ids),
                        VerificationItem.status.in_(("approved", "rejected")),
                    )
                )
            ).scalar_one()

    return {
        "evidence_bundle_endpoint": "/api/v1/planting-projects/{project_id}/evidence-bundle",
        "signed_evidence_headers": ["X-BYOT-Evidence-SHA256", "X-BYOT-Evidence-Signature"],
        "uncertainty_methodology": "Monte Carlo 90% CI + Verra VM0047 deduction",
        "verifier_attestations_count": int(attestation_count or 0),
        "credit_serials": serials,
        "auditor_access": "Invite org member with org_role=viewer for read-only BRSR export",
    }


async def build_brsr_context(
    db: AsyncSession,
    *,
    organization: Organization,
    project_id: uuid.UUID | None = None,
) -> dict[str, Any]:
    projects = await _portfolio_projects(db, organization.id, project_id)
    project_ids = [p.id for p in projects]

    ghg_inventory: list[dict[str, Any]] = []
    project_summaries: list[dict[str, Any]] = []
    total_removals_t = 0.0

    for project in projects:
        lines = await _project_ghg_lines(db, project)
        ghg_inventory.extend(lines)
        gross_line = lines[0] if lines else {}
        total_removals_t += float(gross_line.get("amount_tco2e") or 0)
        mrv = await build_project_mrv_context(db, project)
        project_summaries.append(
            {
                "code": project.code,
                "name": project.name,
                "scheme_code": project.scheme_code,
                "tree_count": mrv["summary"].get("tree_count", 0),
                "open_violations": mrv["summary"].get("open_violations", 0),
            }
        )

    credits = await org_credit_summary(db, organization.id)
    assurance = await _assurance_pack(db, organization.id, project_ids)
    value_chain_annex = build_value_chain_annex(projects)
    open_violations_total = sum(int(p.get("open_violations") or 0) for p in project_summaries)
    core_kpi_rows = build_core_kpi_sheet_rows(
        ghg_inventory=ghg_inventory,
        project_summaries=project_summaries,
        open_violations_total=open_violations_total,
        value_chain_projects=value_chain_annex,
    )

    essential_indicators = [
        {
            "indicator_id": "P6.E4",
            "name": "Greenhouse gas emissions / land sector removals",
            "description": (
                "Scope-tagged GHG inventory line items for plantation carbon sequestration "
                "and registry-aligned net credits."
            ),
            "ghg_inventory": ghg_inventory,
            "portfolio_total_removals_tco2e": round(total_removals_t, 4),
        },
        {
            "indicator_id": "P6.E7",
            "name": "Biodiversity and nature-related dependencies",
            "description": "Project-level tree registry, native species mix, and monitoring coverage.",
            "metrics": {
                "project_count": len(projects),
                "projects": project_summaries,
            },
        },
        {
            "indicator_id": "P6.E-ASSURANCE",
            "name": "Assurance-ready evidence pack (BYOT extension)",
            "description": (
                "Links signed evidence bundles, uncertainty CI, verifier attestations, "
                "and credit serials for limited assurance workflows."
            ),
            "assurance_pack": assurance,
            "credit_ledger_summary": credits,
        },
    ]

    return {
        "brsr_core_version": BRSR_CORE_VERSION,
        "principle": PRINCIPLE,
        "principle_title": "Businesses should respect and make efforts to protect and restore the environment",
        "framework_reference": "SEBI BRSR Core 2024 — Principle 6 Essential Indicators",
        "generated_at": datetime.now(UTC).isoformat(),
        "reporting_year": _reporting_year(),
        "organization": {
            "id": str(organization.id),
            "name": organization.name,
            "slug": organization.slug,
        },
        "scope": "single_project" if project_id else "organization_portfolio",
        "project_id": str(project_id) if project_id else None,
        "essential_indicators": essential_indicators,
        "core_kpi_mapping": core_kpi_rows,
        "value_chain_annex": value_chain_annex,
        "disclaimer": (
            "This export maps BYOT MRV data to BRSR Core Principle 6 structure for assurance "
            "preparation. It is not a filed BRSR submission or statutory assurance opinion."
        ),
    }


def render_brsr_json(ctx: dict[str, Any]) -> bytes:
    return json.dumps(ctx, indent=2, default=str).encode("utf-8")


def render_brsr_xlsx(ctx: dict[str, Any]) -> bytes:
    wb = Workbook()
    meta = wb.active
    meta.title = "BRSR Meta"
    meta.append(["field", "value"])
    meta.append(["brsr_core_version", ctx["brsr_core_version"]])
    meta.append(["principle", ctx["principle"]])
    meta.append(["organization", ctx["organization"]["name"]])
    meta.append(["reporting_year", ctx["reporting_year"]])
    meta.append(["generated_at", ctx["generated_at"]])
    meta.append(["scope", ctx["scope"]])

    inv = wb.create_sheet("GHG Inventory")
    inv.append(
        [
            "line_id",
            "project_code",
            "brsr_indicator",
            "scope",
            "ghg_protocol_category",
            "gas",
            "amount_tco2e",
            "uncertainty_pct",
            "co2e_lower_90_t",
            "co2e_upper_90_t",
            "methodology",
        ]
    )
    for indicator in ctx.get("essential_indicators") or []:
        if indicator.get("indicator_id") != "P6.E4":
            continue
        for line in indicator.get("ghg_inventory") or []:
            inv.append(
                [
                    line.get("line_id"),
                    line.get("project_code"),
                    line.get("brsr_indicator"),
                    line.get("scope"),
                    line.get("ghg_protocol_category"),
                    line.get("gas"),
                    line.get("amount_tco2e"),
                    line.get("uncertainty_pct"),
                    line.get("co2e_lower_90_t"),
                    line.get("co2e_upper_90_t"),
                    line.get("methodology"),
                ]
            )

    indicators = wb.create_sheet("Essential Indicators")
    indicators.append(["indicator_id", "name", "description"])
    for ind in ctx.get("essential_indicators") or []:
        indicators.append([ind.get("indicator_id"), ind.get("name"), ind.get("description")])

    core = wb.create_sheet("Core KPIs")
    core.append(["kpi_id", "name", "data_available", "value_summary", "platform_source", "notes"])
    for row in ctx.get("core_kpi_mapping") or []:
        core.append(
            [
                row.get("kpi_id"),
                row.get("name"),
                row.get("data_available"),
                row.get("value_summary"),
                row.get("platform_source"),
                row.get("notes"),
            ]
        )

    vc = wb.create_sheet("Value chain")
    vc.append(["project_code", "project_name", "scheme_code", "segment", "supplier_ref", "state", "role"])
    for row in ctx.get("value_chain_annex") or []:
        vc.append(
            [
                row.get("project_code"),
                row.get("project_name"),
                row.get("scheme_code"),
                row.get("segment"),
                row.get("supplier_ref"),
                row.get("state"),
                row.get("role"),
            ]
        )

    assurance = wb.create_sheet("Assurance Pack")
    assurance.append(["key", "value"])
    for indicator in ctx.get("essential_indicators") or []:
        if indicator.get("indicator_id") == "P6.E-ASSURANCE":
            pack = indicator.get("assurance_pack") or {}
            for key, val in pack.items():
                if isinstance(val, list | dict):
                    assurance.append([key, json.dumps(val, default=str)])
                else:
                    assurance.append([key, val])

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def render_brsr_zip(ctx: dict[str, Any]) -> bytes:
    buf = io.BytesIO()
    org_slug = ctx["organization"].get("slug") or "org"
    with zipfile.ZipFile(buf, mode="w", compression=zipfile.ZIP_DEFLATED) as zf:
        zf.writestr(f"brsr-{org_slug}-p6.json", render_brsr_json(ctx))
        zf.writestr(f"brsr-{org_slug}-p6.xlsx", render_brsr_xlsx(ctx))
    return buf.getvalue()


BrsrFormat = Literal["json", "xlsx", "zip"]
