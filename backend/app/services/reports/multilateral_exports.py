"""Multilateral & DFI compliance exports (Phase C — World Bank ESF, UNDP SES)."""

from __future__ import annotations

import io
import json
import zipfile
from datetime import UTC, datetime
from typing import Any

from openpyxl import Workbook
from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession

from app.models.plantation_fence import PlantationFence
from app.models.planting_project import PlantingProject
from app.services.compliance.safeguards import list_safeguard_documents
from app.services.planting_projects.mrv_export import build_project_mrv_context
from app.services.reports.carbon_integrity_context import build_carbon_integrity_envelope
from app.services.reports.india_exports import (
    build_campa_state_export_context,
    render_campa_state_export_xlsx,
)
from app.services.reports.tnfd import _fence_bioacoustic_summary, _fence_ndvi_summary

DISCLAIMER = (
    "Prepared for multilateral and development-finance audit review. "
    "Not World Bank, UNDP, or lender certification."
)


async def _site_biodiversity_rows(db: AsyncSession, project_id: Any) -> list[dict[str, Any]]:
    fences = list(
        (
            await db.execute(
                select(PlantationFence)
                .where(PlantationFence.project_id == project_id)
                .order_by(PlantationFence.name.asc())
            )
        ).scalars().all()
    )
    rows: list[dict[str, Any]] = []
    for fence in fences:
        ndvi = await _fence_ndvi_summary(db, fence.id)
        bio = await _fence_bioacoustic_summary(db, fence.id)
        rows.append(
            {
                "site_name": fence.name,
                "fence_id": str(fence.id),
                "ndvi_mean": ndvi.get("ndvi_mean"),
                "ndvi_scene_date": ndvi.get("scene_acquired_at"),
                "satellite_provider": ndvi.get("provider"),
                "species_richness": bio.get("species_richness", 0),
                "recording_count": bio.get("recording_count", 0),
                "threatened_signals": bio.get("threatened_signals", 0),
                "avg_shannon_diversity": bio.get("avg_shannon_diversity"),
            }
        )
    return rows


def _ses_risk_tier(
    *,
    open_violations: int,
    safeguard_doc_count: int,
    native_species_pct: float | None,
    biodiversity_sites: int,
) -> str:
    if open_violations > 0:
        return "high"
    if safeguard_doc_count < 2:
        return "high"
    if safeguard_doc_count < 4 or (native_species_pct is not None and native_species_pct < 50):
        return "medium"
    if biodiversity_sites == 0:
        return "medium"
    return "low"


async def build_esf_ps5_context(db: AsyncSession, project: PlantingProject) -> dict[str, Any]:
    mrv = await build_project_mrv_context(db, project)
    safeguards = await list_safeguard_documents(db, project)
    summary = mrv.get("summary") or {}
    return {
        "export_type": "world_bank_esf_ps5",
        "generated_at": datetime.now(UTC).isoformat(),
        "disclaimer": DISCLAIMER,
        "framework_reference": "World Bank Environmental and Social Framework — PS5 Land & Tenure",
        "project": {
            "code": project.code,
            "name": project.name,
            "scheme_code": project.scheme_code,
        },
        "tenure_evidence": {
            "safeguard_documents": safeguards,
            "document_count": len(safeguards),
            "doc_types_present": sorted({d["doc_type"] for d in safeguards}),
        },
        "compliance": {
            "open_violations": int(summary.get("open_violations") or 0),
            "work_area_count": int(summary.get("work_area_count") or 0),
        },
    }


async def build_esf_ps6_context(db: AsyncSession, project: PlantingProject) -> dict[str, Any]:
    mrv = await build_project_mrv_context(db, project)
    summary = mrv.get("summary") or {}
    sites = await _site_biodiversity_rows(db, project.id)
    return {
        "export_type": "world_bank_esf_ps6",
        "generated_at": datetime.now(UTC).isoformat(),
        "disclaimer": DISCLAIMER,
        "framework_reference": "World Bank ESF — PS6 Biodiversity Conservation",
        "project": {
            "code": project.code,
            "name": project.name,
            "scheme_code": project.scheme_code,
        },
        "biodiversity": {
            "native_species_pct": summary.get("native_species_pct"),
            "tree_count": int(summary.get("tree_count") or 0),
            "sites": sites,
            "sites_with_bioacoustic": sum(1 for s in sites if s["recording_count"] > 0),
            "sites_with_ndvi": sum(1 for s in sites if s["ndvi_mean"] is not None),
        },
    }


async def build_undp_ses_context(db: AsyncSession, project: PlantingProject) -> dict[str, Any]:
    mrv = await build_project_mrv_context(db, project)
    summary = mrv.get("summary") or {}
    safeguards = await list_safeguard_documents(db, project)
    sites = await _site_biodiversity_rows(db, project.id)
    stakeholder_docs = [d for d in safeguards if d["doc_type"] == "stakeholder_consultation_log"]
    risk_tier = _ses_risk_tier(
        open_violations=int(summary.get("open_violations") or 0),
        safeguard_doc_count=len(safeguards),
        native_species_pct=summary.get("native_species_pct"),
        biodiversity_sites=sum(1 for s in sites if s["recording_count"] > 0 or s["ndvi_mean"]),
    )
    mitigations: list[str] = []
    if int(summary.get("open_violations") or 0) > 0:
        mitigations.append("Resolve open compliance violations before lender disbursement.")
    if len(safeguards) < 4:
        mitigations.append("Complete safeguards document set (FPIC, tenure, stakeholder log).")
    if not sites:
        mitigations.append("Register work areas and run satellite / bioacoustic monitoring.")
    if risk_tier == "low":
        mitigations.append("Maintain periodic stakeholder engagement and survival monitoring.")

    return {
        "export_type": "undp_ses_screening",
        "generated_at": datetime.now(UTC).isoformat(),
        "disclaimer": DISCLAIMER,
        "framework_reference": "UNDP Social and Environmental Standards (SES)",
        "project": {
            "code": project.code,
            "name": project.name,
            "scheme_code": project.scheme_code,
        },
        "screening": {
            "risk_tier": risk_tier,
            "required_mitigations": mitigations,
        },
        "stakeholder_engagement": {
            "log_entries": stakeholder_docs,
            "total_safeguard_documents": len(safeguards),
            "all_documents": safeguards,
        },
    }


def _write_sheet_from_rows(ws, headers: list[str], rows: list[list[Any]]) -> None:
    ws.append(headers)
    for row in rows:
        ws.append(row)


def render_esf_ps5_xlsx(ctx: dict[str, Any]) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "ESF PS5"
    project = ctx.get("project") or {}
    tenure = ctx.get("tenure_evidence") or {}
    ws.append(["World Bank ESF — PS5 Land & Tenure Evidence"])
    ws.append(["Disclaimer", ctx.get("disclaimer", "")])
    ws.append(["Project", f"{project.get('code', '')} — {project.get('name', '')}"])
    ws.append(["Documents on file", tenure.get("document_count", 0)])
    ws.append([])
    ws.append(["Doc type", "Title", "Uploaded", "S3 key"])
    for doc in tenure.get("safeguard_documents") or []:
        ws.append(
            [
                doc.get("doc_type_label") or doc.get("doc_type"),
                doc.get("title"),
                doc.get("created_at"),
                doc.get("s3_key"),
            ]
        )
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def render_esf_ps6_xlsx(ctx: dict[str, Any]) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "ESF PS6"
    bio = ctx.get("biodiversity") or {}
    project = ctx.get("project") or {}
    ws.append(["World Bank ESF — PS6 Biodiversity Evidence"])
    ws.append(["Disclaimer", ctx.get("disclaimer", "")])
    ws.append(["Project", f"{project.get('code', '')} — {project.get('name', '')}"])
    ws.append(["Native species %", bio.get("native_species_pct", "—")])
    ws.append(["Trees registered", bio.get("tree_count", 0)])
    ws.append([])
    _write_sheet_from_rows(
        ws,
        [
            "Site",
            "NDVI mean",
            "NDVI scene date",
            "Species richness",
            "Recordings",
            "Threatened signals",
            "Shannon diversity",
        ],
        [
            [
                s.get("site_name"),
                s.get("ndvi_mean"),
                s.get("ndvi_scene_date"),
                s.get("species_richness"),
                s.get("recording_count"),
                s.get("threatened_signals"),
                s.get("avg_shannon_diversity"),
            ]
            for s in bio.get("sites") or []
        ],
    )
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def render_undp_ses_xlsx(ctx: dict[str, Any]) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "UNDP SES"
    screening = ctx.get("screening") or {}
    stakeholder = ctx.get("stakeholder_engagement") or {}
    project = ctx.get("project") or {}
    ws.append(["UNDP SES — Risk Screening"])
    ws.append(["Disclaimer", ctx.get("disclaimer", "")])
    ws.append(["Project", f"{project.get('code', '')} — {project.get('name', '')}"])
    ws.append(["Risk tier", screening.get("risk_tier", "—")])
    ws.append([])
    ws.append(["Required mitigations"])
    for item in screening.get("required_mitigations") or []:
        ws.append(["", item])
    ws.append([])
    ws.append(["Stakeholder engagement log"])
    ws.append(["Doc type", "Title", "Uploaded"])
    for doc in stakeholder.get("log_entries") or []:
        ws.append([doc.get("doc_type_label"), doc.get("title"), doc.get("created_at")])
    ws.append([])
    ws.append(["All safeguard documents"])
    ws.append(["Doc type", "Title", "Uploaded"])
    for doc in stakeholder.get("all_documents") or []:
        ws.append([doc.get("doc_type_label"), doc.get("title"), doc.get("created_at")])
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


async def build_multilateral_audit_pack(
    db: AsyncSession,
    project: PlantingProject,
) -> tuple[bytes, dict[str, Any]]:
    """Combined ZIP: ESF PS5/PS6, UNDP SES, MRV context, carbon integrity, optional CAMPA."""
    ps5_ctx = await build_esf_ps5_context(db, project)
    ps6_ctx = await build_esf_ps6_context(db, project)
    ses_ctx = await build_undp_ses_context(db, project)
    mrv = await build_project_mrv_context(db, project)
    integrity = await build_carbon_integrity_envelope(db, project)

    manifest_files: list[dict[str, str]] = []
    buf = io.BytesIO()

    def add(path: str, data: bytes) -> None:
        manifest_files.append({"path": path, "size_bytes": str(len(data))})

    with zipfile.ZipFile(buf, mode="w", compression=zipfile.ZIP_DEFLATED) as zf:
        readme = (
            "Aranyix Multilateral Audit Pack\n"
            "================================\n\n"
            "Combined ESF PS5/PS6, UNDP SES screening, MRV context, and carbon integrity.\n"
            "For NHAI, CAMPA, and DFI-backed green corridor audits.\n\n"
            f"{DISCLAIMER}\n"
        )
        zf.writestr("README.txt", readme)
        add("README.txt", readme.encode())

        ps5_xlsx = render_esf_ps5_xlsx(ps5_ctx)
        zf.writestr("esf-ps5-tenure-evidence.xlsx", ps5_xlsx)
        add("esf-ps5-tenure-evidence.xlsx", ps5_xlsx)

        ps6_xlsx = render_esf_ps6_xlsx(ps6_ctx)
        zf.writestr("esf-ps6-biodiversity-evidence.xlsx", ps6_xlsx)
        add("esf-ps6-biodiversity-evidence.xlsx", ps6_xlsx)

        ses_xlsx = render_undp_ses_xlsx(ses_ctx)
        zf.writestr("undp-ses-screening.xlsx", ses_xlsx)
        add("undp-ses-screening.xlsx", ses_xlsx)

        mrv_json = json.dumps(mrv, indent=2, default=str).encode()
        zf.writestr("mrv-context.json", mrv_json)
        add("mrv-context.json", mrv_json)

        integrity_json = json.dumps(integrity, indent=2, default=str).encode()
        zf.writestr("carbon-integrity.json", integrity_json)
        add("carbon-integrity.json", integrity_json)

        if project.scheme_code in {"campa_ca", "nhai_highway", "dfi_green_corridor"}:
            campa_ctx = await build_campa_state_export_context(db, project)
            campa_xlsx = render_campa_state_export_xlsx(campa_ctx)
            zf.writestr("state-campa-monitoring.xlsx", campa_xlsx)
            add("state-campa-monitoring.xlsx", campa_xlsx)

        pack_manifest = {
            "pack_version": "aranyix-multilateral-1.0.0",
            "project_code": project.code,
            "generated_at": datetime.now(UTC).isoformat(),
            "ses_risk_tier": ses_ctx["screening"]["risk_tier"],
            "files": manifest_files,
        }
        manifest_json = json.dumps(pack_manifest, indent=2).encode()
        zf.writestr("manifest.json", manifest_json)

    zip_bytes = buf.getvalue()
    summary = {
        "project_code": project.code,
        "file_count": len(manifest_files) + 1,
        "ses_risk_tier": ses_ctx["screening"]["risk_tier"],
        "zip_size_bytes": len(zip_bytes),
    }
    return zip_bytes, summary
