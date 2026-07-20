"""PDF and Excel renderers for framework-mapped reports."""

from __future__ import annotations

import io
from datetime import datetime
from typing import Any

from openpyxl import Workbook
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.lib.units import mm
from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle


def _table_style_header() -> TableStyle:
    return TableStyle(
        [
            ("BACKGROUND", (0, 0), (0, -1), colors.HexColor("#DCFCE7")),
            ("TEXTCOLOR", (0, 0), (0, -1), colors.HexColor("#15803D")),
            ("GRID", (0, 0), (-1, -1), 0.25, colors.grey),
            ("FONTNAME", (0, 0), (-1, -1), "Helvetica"),
            ("FONTSIZE", (0, 0), (-1, -1), 9),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 5),
        ]
    )


def render_framework_report_pdf(ctx: dict[str, Any]) -> bytes:
    framework = ctx["framework"]
    project = ctx["project"]
    buf = io.BytesIO()
    doc = SimpleDocTemplate(
        buf,
        pagesize=A4,
        title=f"{framework['short_label']} — {project['name']}",
        author="Aranyix BYOT",
        leftMargin=18 * mm,
        rightMargin=18 * mm,
        topMargin=18 * mm,
        bottomMargin=18 * mm,
    )
    styles = getSampleStyleSheet()
    h1 = styles["Heading1"]
    h1.textColor = colors.HexColor("#15803D")
    h2 = styles["Heading2"]
    body = styles["BodyText"]
    small = styles["BodyText"]
    small.fontSize = 8
    small.textColor = colors.grey

    story: list = []
    story.append(Paragraph(framework["title"], h1))
    story.append(
        Paragraph(
            f"<b>{project['name']}</b> ({project['code']}) · "
            f"profile: <b>{framework['code']}</b> · "
            f"generated {datetime.utcnow().strftime('%Y-%m-%d %H:%M UTC')}",
            body,
        )
    )
    story.append(Paragraph(f"Reference: {framework['reference']}", small))
    story.append(Spacer(1, 4 * mm))
    story.append(
        Paragraph(
            f"<i>{framework['disclaimer']}</i>",
            small,
        )
    )
    story.append(Spacer(1, 6 * mm))

    carbon = ctx.get("carbon_summary") or {}
    if carbon.get("methodology") and carbon["methodology"] != "NONE":
        story.append(Paragraph("Carbon summary", h2))
        carbon_rows = [
            ["Methodology", carbon.get("methodology", "—")],
            ["Engine version", carbon.get("engine_version", "—")],
            ["Total CO₂e (kg)", f"{carbon.get('total_co2e_kg', 0):,.2f}"],
            ["Gross credits (tCO₂e)", f"{carbon.get('gross_credits_tco2e', 0):.4f}"],
        ]
        if carbon.get("buffer_pct", 0) > 0:
            carbon_rows.append(
                ["Buffer withheld", f"{carbon.get('buffer_withheld_tco2e', 0):.4f} tCO₂e"]
            )
            carbon_rows.append(
                ["Net credits estimate", f"{carbon.get('net_credits_tco2e', 0):.4f} tCO₂e"]
            )
        t = Table(carbon_rows, colWidths=[80 * mm, 80 * mm])
        t.setStyle(_table_style_header())
        story.append(t)
        story.append(Spacer(1, 6 * mm))

    for section in ctx.get("sections") or []:
        story.append(Paragraph(section.get("title", "Section"), h2))
        rows = section.get("rows") or []
        if rows:
            t = Table(rows, colWidths=[80 * mm, 80 * mm])
            t.setStyle(_table_style_header())
            story.append(t)
        story.append(Spacer(1, 5 * mm))

    story.append(Paragraph("Project KPIs", h2))
    summary = ctx.get("summary") or {}
    kpi_rows = [
        ["Trees registered", str(summary.get("tree_count", 0))],
        ["Work areas", str(summary.get("work_area_count", 0))],
        ["Open violations", str(summary.get("open_violations", 0))],
        ["Native species %", str(summary.get("native_species_pct") or "—")],
    ]
    story.append(Table(kpi_rows, colWidths=[80 * mm, 80 * mm]))
    story.append(Spacer(1, 6 * mm))

    story.append(Paragraph("Tree registry (sample)", h2))
    tree_table = [["Code", "Species", "Health", "Survival"]]
    for tr in (ctx.get("trees") or [])[:30]:
        tree_table.append(
            [
                tr.get("public_code", ""),
                tr.get("species", ""),
                tr.get("health", ""),
                tr.get("survival_status", ""),
            ]
        )
    if len(tree_table) == 1:
        tree_table.append(["—", "—", "—", "—"])
    dt = Table(tree_table, repeatRows=1)
    dt.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#15803D")),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                ("FONTNAME", (0, 0), (-1, -1), "Helvetica"),
                ("FONTSIZE", (0, 0), (-1, -1), 8),
                ("GRID", (0, 0), (-1, -1), 0.1, colors.grey),
            ]
        )
    )
    story.append(dt)

    doc.build(story)
    return buf.getvalue()


def render_framework_report_xlsx(ctx: dict[str, Any]) -> bytes:
    wb = Workbook()
    framework = ctx["framework"]
    project = ctx["project"]

    ws = wb.active
    ws.title = "Summary"
    ws.append(["field", "value"])
    ws.append(["framework", framework["code"]])
    ws.append(["title", framework["title"]])
    ws.append(["project", project["name"]])
    ws.append(["project_code", project["code"]])
    ws.append(["reference", framework["reference"]])
    ws.append(["disclaimer", framework["disclaimer"]])
    ws.append([])
    for key, val in (ctx.get("carbon_summary") or {}).items():
        ws.append([key, val])

    for idx, section in enumerate(ctx.get("sections") or []):
        title = (section.get("title") or f"Section {idx + 1}")[:31]
        sh = wb.create_sheet(title=title)
        sh.append(["metric", "value"])
        for row in section.get("rows") or []:
            if len(row) >= 2:
                sh.append([row[0], row[1]])

    trees_ws = wb.create_sheet(title="Trees")
    trees_ws.append(["public_code", "species", "health", "survival", "lat", "lon"])
    for tr in ctx.get("trees") or []:
        trees_ws.append(
            [
                tr.get("public_code"),
                tr.get("species"),
                tr.get("health"),
                tr.get("survival_status"),
                tr.get("lat"),
                tr.get("lon"),
            ]
        )

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
